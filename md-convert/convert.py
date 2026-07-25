#!/usr/bin/env python3
"""md-convert — turn a Markdown file into .docx / .html / .pdf / .xlsx.

Part of the global `md-convert` skill (~/.claude/skills/md-convert/). Zero
assumptions about any repo: uses pandoc for docx/html/pdf and stdlib-ish
parsing for xlsx (tables + task lists → sheets; needs openpyxl available in
whichever python runs this — pass --python or run with a venv python).

Usage:
    python3 convert.py INPUT.md --to docx [--out PATH] [--stamp | --no-stamp]

Datetime stamping (--stamp): prepends a small "Generated: <local datetime>"
line to the output (docx/pdf/html: a subtitle-style first line; xlsx: cell A1
of a "meta" sheet). Because these files get overwritten often, the stamp shows
readers which vintage they hold. The DEFAULT comes from config.json next to
this script ({"stamp_default": true|false}); explicit flags always win.
"""

from __future__ import annotations

import argparse
import datetime
import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

SKILL_DIR = Path(__file__).resolve().parent
CONFIG = SKILL_DIR / "config.json"
FORMATS = ("docx", "html", "pdf", "xlsx")


def stamp_default() -> bool | None:
    """None means unset — the skill instructions tell Claude to ask the user."""
    try:
        return json.loads(CONFIG.read_text()).get("stamp_default")
    except (OSError, ValueError):
        return None


def _stamped_md(src: Path, stamp: bool) -> Path:
    if not stamp:
        return src
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M %Z").strip()
    body = src.read_text()
    # place the stamp AFTER a leading H1 so titles stay first; else prepend
    lines = body.splitlines()
    stamp_line = f"*Generated: {now}*"
    if lines and lines[0].startswith("# "):
        lines.insert(1, "")
        lines.insert(2, stamp_line)
    else:
        lines = [stamp_line, ""] + lines
    tmp = Path(tempfile.mkstemp(suffix=".md")[1])
    tmp.write_text("\n".join(lines))
    return tmp


def _pandoc(src: Path, out: Path, fmt: str, resource_dir: Path | None = None) -> None:
    # -blank_before_header: GitHub-style leniency — recognize '# Heading'
    # even without a preceding blank line (pandoc's default md dialect
    # silently swallows such headings into the previous paragraph).
    cmd = ["pandoc", "-f", "markdown-blank_before_header",
           str(src), "-o", str(out), "--standalone"]
    # Resolve relative image paths (e.g. ![](figures/F2.png)) against the
    # ORIGINAL markdown's directory, so figures embed even when --stamp copied
    # the md to a temp dir. For docx/pptx/pdf pandoc embeds found images
    # automatically; it just needs to find them.
    if resource_dir is not None:
        cmd += [f"--resource-path={resource_dir}"]
    if fmt == "pdf":
        # prefer weasyprint/wkhtmltopdf if present; else LaTeX if present
        for engine in ("weasyprint", "wkhtmltopdf", "pdflatex", "tectonic"):
            if shutil.which(engine):
                cmd += [f"--pdf-engine={engine}"]
                break
        else:
            sys.exit(
                "pdf needs an engine — install one of: brew install weasyprint | "
                "brew install --cask wkhtmltopdf | brew install tectonic"
            )
    subprocess.run(cmd, check=True)


_TOC_STYLE = (
    '<w:style w:type="paragraph" w:styleId="TOC{n}">'
    '<w:name w:val="toc {n}"/><w:basedOn w:val="Normal"/>'
    '<w:pPr><w:ind w:left="{ind}"/><w:spacing w:after="60"/></w:pPr>'
    "</w:style>"
)


def _add_word_toc(out: Path, depth: int = 3) -> None:
    """
    Post-process a pandoc-produced .docx: give it a Word-native table of
    contents (proven on the llm-as-judge manuscript review loop, 2026-07).

    - Every Heading 1..depth paragraph gets a _Toc bookmark.
    - TOC entries are TOC1/2/3-styled paragraphs hyperlinked to those
      bookmarks (Ctrl/Cmd+click jumps to the section), pre-cached so the
      TOC is visible immediately on open, inside a live TOC field
      (right-click -> Update Field in Word regenerates it + page numbers).
    - A lone leading Heading 1 is treated as the document title: excluded
      from the TOC, which is inserted before the first indexed heading.
    - Heading levels are normalized (shallowest indexed level -> TOC1).

    Robustness note: heading discovery reads the *rendered* Heading styles
    in document.xml (so '#' lines inside code blocks can't pollute the
    TOC), but the paragraph regex is coupled to pandoc's docx XML shape —
    the pairing assertions below fail loudly if a pandoc upgrade changes it.
    """
    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        parts = {n: z.read(n) for n in names}
    xml = parts["word/document.xml"].decode("utf8")
    styles = parts["word/styles.xml"].decode("utf8")

    heads = []  # (level, text, span_start, span_end, pPr_end_offset)
    for m in re.finditer(
        rf'<w:p><w:pPr><w:pStyle w:val="Heading([1-{depth}])"\s*/>'
        r"(?:(?!</w:p>).)*?</w:p>", xml, flags=re.S,
    ):
        text = "".join(re.findall(r"<w:t[^>]*>([^<]*)</w:t>", m.group(0)))
        heads.append([int(m.group(1)), text, m.start(), m.end()])
    if not heads:
        sys.exit("--toc: no Heading 1-3 paragraphs found in the converted docx")

    # Lone leading H1 = document title, not a section.
    if heads[0][0] == 1 and sum(1 for h in heads if h[0] == 1) == 1:
        heads = heads[1:]
    if not heads:
        sys.exit("--toc: only a title heading found; nothing to index")
    base = min(h[0] for h in heads)

    entries = []  # (toc_level, text, bookmark_name)
    offset_shift = 0
    for n, (level, text, s, e) in enumerate(heads):
        name = f"_Toc9{n:07d}"
        bm = (f'<w:bookmarkStart w:id="{7000 + n}" w:name="{name}"/>'
              f'<w:bookmarkEnd w:id="{7000 + n}"/>')
        s += offset_shift
        e += offset_shift
        para = xml[s:e].replace("</w:pPr>", "</w:pPr>" + bm, 1)
        xml = xml[:s] + para + xml[e:]
        offset_shift += len(bm)
        entries.append((min(level - base + 1, 3), text, name))

    ps = ['<w:p><w:pPr><w:pStyle w:val="TOCHeading"/></w:pPr>'
          "<w:r><w:t>Contents</w:t></w:r></w:p>"]
    for n, (lvl, text, name) in enumerate(entries):
        first = ('<w:r><w:fldChar w:fldCharType="begin"/></w:r>'
                 '<w:r><w:instrText xml:space="preserve"> '
                 'TOC \\o "1-3" \\h \\z \\u </w:instrText></w:r>'
                 '<w:r><w:fldChar w:fldCharType="separate"/></w:r>') if n == 0 else ""
        last = '<w:r><w:fldChar w:fldCharType="end"/></w:r>' if n == len(entries) - 1 else ""
        link = (f'<w:hyperlink w:anchor="{name}" w:history="1">'
                '<w:r><w:rPr><w:rStyle w:val="Hyperlink"/></w:rPr>'
                f'<w:t xml:space="preserve">{escape(text)}</w:t></w:r></w:hyperlink>')
        ps.append(f'<w:p><w:pPr><w:pStyle w:val="TOC{lvl}"/></w:pPr>{first}{link}{last}</w:p>')

    # Insert before the first indexed heading's (now bookmarked) paragraph.
    i_first = xml.find(f'w:name="{entries[0][2]}"')
    i_para = xml.rfind("<w:p>", 0, i_first)
    xml = xml[:i_para] + "".join(ps) + xml[i_para:]

    # TOCHeading exists in pandoc's reference doc; TOC1-3 usually don't.
    add = "".join(
        _TOC_STYLE.format(n=n, ind=(n - 1) * 360)
        for n in (1, 2, 3) if f'w:styleId="TOC{n}"' not in styles
    )
    if add:
        styles = styles.replace("</w:styles>", add + "</w:styles>")

    n_links = xml.count('<w:hyperlink w:anchor="_Toc9')
    n_bm = xml.count('w:name="_Toc9')
    if n_links == 0 or n_links != n_bm:
        sys.exit(f"--toc verification FAILED (hyperlinks={n_links}, bookmarks={n_bm})")

    parts["word/document.xml"] = xml.encode("utf8")
    parts["word/styles.xml"] = styles.encode("utf8")
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for n in names:
            z.writestr(n, parts[n])
    print(f"TOC: {len(entries)} hyperlinked entries (update field in Word for page numbers)")


def _xlsx(src: Path, out: Path, stamp: bool) -> None:
    try:
        import openpyxl
    except ImportError:
        sys.exit("xlsx needs openpyxl — run with a python that has it (pip install openpyxl)")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "content"
    row = 1
    if stamp:
        meta = wb.create_sheet("meta")
        meta["A1"] = f"Generated: {datetime.datetime.now():%Y-%m-%d %H:%M}"
    in_table = False
    for line in src.read_text().splitlines():
        s = line.strip()
        if s.startswith("|") and s.endswith("|"):
            cells = [c.strip() for c in s.strip("|").split("|")]
            if all(set(c) <= {"-", ":", " "} and c for c in cells):
                continue  # separator row
            for col, val in enumerate(cells, start=1):
                ws.cell(row=row, column=col, value=val)
            row += 1
            in_table = True
        elif s:
            if in_table:
                row += 1  # blank spacer after a table
                in_table = False
            ws.cell(row=row, column=1, value=s)
            row += 1
    wb.save(out)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("input", type=Path)
    ap.add_argument("--to", required=True, choices=FORMATS)
    ap.add_argument("--out", type=Path)
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--stamp", dest="stamp", action="store_true", default=None)
    g.add_argument("--no-stamp", dest="stamp", action="store_false")
    ap.add_argument(
        "--toc", action="store_true",
        help="docx only: add a Word-native hyperlinked TOC (headings bookmarked, "
             "entries visible on open, live field — Update Field adds page numbers)",
    )
    args = ap.parse_args()
    if args.toc and args.to != "docx":
        sys.exit("--toc is only supported for --to docx")

    stamp = args.stamp
    if stamp is None:
        stamp = stamp_default()
    if stamp is None:
        sys.exit(
            "stamp default not configured — pass --stamp/--no-stamp, or set it: "
            'echo \'{"stamp_default": true}\' > ' + str(CONFIG)
        )

    out = args.out or args.input.with_suffix(f".{args.to}")
    if args.to == "xlsx":
        _xlsx(args.input, out, stamp)
    else:
        src = _stamped_md(args.input, stamp)
        _pandoc(src, out, args.to, resource_dir=args.input.resolve().parent)
        if args.toc:
            _add_word_toc(out)
    print(f"wrote {out}" + (" (stamped)" if stamp else ""))


if __name__ == "__main__":
    main()
